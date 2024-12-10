from flask import render_template, request, redirect, url_for, flash, Blueprint, Response, jsonify
from app.models import Course
from app import db
from app.models import Lesson, Attendance, Feedback, Course, Lesson
from datetime import datetime, timedelta
from sqlalchemy.sql import func
#for export
import csv
from io import StringIO
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename
import os
#maintain my man

teacher = Blueprint('teacher', __name__, template_folder='templates')
#Overall settings
@teacher.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    from app.models import User
    from app.models import CourseStudent
    teacher_id = current_user.id
    user = User.query.filter_by(id=current_user.id).first()
    courses = Course.query.filter_by(tutor_id=teacher_id).all()  # Получаем все курсы текущего преподавателя

    current_user_data = User.query.get(current_user.id)
    user_name = current_user_data.full_name  # Или другой атрибут, соответствующий имени
    user_role = current_user_data.role  # Или соответствующий атрибут роли
    total_users = User.query.count() #всего пользователей
    total_students = User.query.filter_by(role='student').count() #всего студентов
    total_teachers = User.query.filter_by(role='teacher').count() #всего учителей
    active_users = User.query.filter(User.last_login >= datetime.now() - timedelta(days=1)).count() #активные юзеры за последние 24 часа
    active_courses = Course.query.filter_by(status='active').count() #активные курсы колво

    #самый популярный курс
    # Самый популярный курс
    popular_course_query = (
        db.session.query(
            Course,
            db.func.count(CourseStudent.student_id).label('student_count')
        )
        .join(CourseStudent, Course.id == CourseStudent.course_id)
        .group_by(Course.id)
        .order_by(db.func.count(CourseStudent.student_id).desc())
        .first()
    )
    popular_course = popular_course_query[0].name if popular_course_query else "No courses yet"

    avg_lesson_rating = Feedback.query.with_entities(func.avg(Feedback.mark)).scalar() #средняя оценка по фидбекам
    teacher_feedback_query = Feedback.query.join(Course, Feedback.course_id == Course.id).filter(
        Course.tutor_id == current_user.id
    ).with_entities(func.avg(Feedback.mark)).scalar()
    teacher_rating = round(teacher_feedback_query, 2) if teacher_feedback_query else "N/A"

    return render_template(
        'dashboard.html',
        user = user,
        user_name=current_user_data.full_name,
        user_role = current_user_data.role,
        total_users=total_users,
        total_students=total_students,
        total_teachers=total_teachers,
        active_users=active_users,
        active_courses=active_courses,
        popular_course=popular_course,
        avg_lesson_rating=round(avg_lesson_rating, 2) if avg_lesson_rating else "N/A",
        teacher_rating=teacher_rating,
        courses=courses
    )

@teacher.route('/help', methods=['GET'])
def help_page():
    return render_template('help.html')

@teacher.route('/tips', methods=['GET'])
def tips_page():
    return render_template('tips.html')

@teacher.route('/notifications/get', methods=['GET'])
@login_required
def get_notifications():
    from app.models import Notification
    notifications = Notification.query.filter_by(recipient_id=current_user.id).order_by(Notification.date_time.desc()).all()
    notifications_data = [
        {
            "id": n.id,
            "message": n.message,
            "status": n.status,
            "date_time": n.date_time.strftime('%Y-%m-%d %H:%M:%S'),
            "type": n.type,
        }
        for n in notifications
    ]
    return jsonify(notifications_data)

@teacher.route('/notifications/mark_as_read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_as_read(notification_id):
    from app.models import Notification
    notification = Notification.query.filter_by(id=notification_id, recipient_id=current_user.id).first_or_404()
    notification.status = 'read'
    db.session.commit()
    return jsonify({'success': True})

@teacher.route('/notifications/all', methods=['GET'])
@login_required
def all_notifications():
    from app.models import Notification
    # Получаем все уведомления для текущего пользователя
    notifications = Notification.query.filter_by(recipient_id=current_user.id).order_by(Notification.date_time.desc()).all()
    return render_template('teacher/notifications.html', notifications=notifications)

@teacher.route('/teacher/profile', methods=['GET', 'POST'])
@login_required
def profile():
    from app.models import User
    user = User.query.filter_by(id=current_user.id).first()

    if request.method == 'POST':
        # Собираем обязательные данные из формы
        user.full_name = request.form.get('full_name', user.full_name).strip()
        user.email = request.form.get('email', user.email).strip()
        user.phone_number = request.form.get('phone_number', user.phone_number).strip()

        # Проверка обязательных полей
        if not user.full_name or not user.email or not user.phone_number:
            flash('Full Name, Email, and Phone Number are required fields.', 'error')
            return redirect(url_for('teacher.profile'))

        # Дополнительные необязательные данные
        user.nickname = request.form.get('nickname')
        user.university = request.form.get('university')
        user.institute = request.form.get('institute')
        user.date_of_birth = request.form.get('date_of_birth') or None
        user.address = request.form.get('address')
        user.telegram_link = request.form.get('telegram_link')
        user.two_factor_enabled = bool(request.form.get('two_factor_enabled'))
        user.bio = request.form.get('bio')
        user.favorite_sport_club = request.form.get('favorite_sport_club')

        # Обработка загруженного фото
        if 'profile_picture' in request.files:
            picture = request.files['profile_picture']
            if picture:
                picture_filename = secure_filename(picture.filename)
                picture_path = os.path.join('static', 'images', picture_filename)
                picture.save(picture_path)
                user.profile_picture = picture_path

        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('teacher.profile'))

    return render_template('teacher/profile.html', user=current_user)

@teacher.route('/teacher/update_profile', methods=['POST'])
@login_required
def update_profile():
    from app.models import User
    user = User.query.filter_by(id=current_user.id).first()

    # Собираем данные из формы (обновление)
    user.full_name = request.form.get('full_name', user.full_name).strip()
    user.email = request.form.get('email', user.email).strip()
    user.phone_number = request.form.get('phone_number', user.phone_number).strip()

    # Дополнительные данные
    user.nickname = request.form.get('nickname')
    user.university = request.form.get('university')
    user.institute = request.form.get('institute')
    user.date_of_birth = request.form.get('date_of_birth') or None
    user.address = request.form.get('address')
    user.telegram_link = request.form.get('telegram_link')
    user.two_factor_enabled = bool(request.form.get('two_factor_enabled'))
    user.bio = request.form.get('bio')
    user.favorite_sport_club = request.form.get('favorite_sport_club')

    # Обработка изображения профиля
    if 'profile_picture' in request.files:
        picture = request.files['profile_picture']
        if picture:
            picture_filename = secure_filename(picture.filename)
            picture_path = os.path.join('static', 'images', picture_filename)
            picture.save(picture_path)
            user.profile_picture = picture_path

    db.session.commit()
    flash('Profile updated successfully!', 'success')
    return redirect(url_for('teacher.profile'))

#Courses
@teacher.route('/courses', methods=['GET'])
def courses():
    teacher_id = current_user.id
    courses = Course.query.filter_by(tutor_id=teacher_id).all()
    return render_template('teacher/courses.html', courses=courses)

@teacher.route('/courses/add', methods=['GET', 'POST'])
def add_course():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        course_code = request.form.get('course_code')
        category = request.form.get('category')

        # Заглушка для ID преподавателя
        teacher_id = current_user.id

        new_course = Course(
            name=name,
            description=description,
            tutor_id=teacher_id,
            start_date=start_date,
            end_date=end_date,
            course_code=course_code,
            category=category
        )
        db.session.add(new_course)
        db.session.commit()
        flash('Course added successfully!', category='success')
        return redirect(url_for('teacher.courses'))
    return render_template('teacher/add_course.html')

@teacher.route('/courses/<int:course_id>', methods=['GET'])
@login_required
def course_details(course_id):
    course = Course.query.filter_by(id=course_id, tutor_id=current_user.id).first_or_404()
    lessons = Lesson.query.filter_by(course_id=course_id).order_by(Lesson.date).all()
    total_lessons = len(lessons)
    completed_lessons = sum(1 for lesson in lessons if lesson.date <= datetime.utcnow().date())
    remaining_lessons = total_lessons - completed_lessons

    return render_template(
        'teacher/course_details.html',
        course=course,
        lessons=lessons,
        statistics={
            'total_lessons': total_lessons,
            'completed_lessons': completed_lessons,
            'remaining_lessons': remaining_lessons
        }
    )

@teacher.route('/courses/<int:course_id>/edit', methods=['GET', 'POST'])
def edit_course(course_id):
    course = Course.query.get_or_404(course_id)
    if request.method == 'POST':
        course.name = request.form.get('name')
        course.description = request.form.get('description')
        course.start_date = request.form.get('start_date')
        course.end_date = request.form.get('end_date')
        course.course_code = request.form.get('course_code')
        course.category = request.form.get('category')
        db.session.commit()
        flash('Course updated successfully!', category='success')
        return redirect(url_for('teacher.courses'))
    return render_template('teacher/edit_course.html', course=course)

@teacher.route('/courses/<int:course_id>/delete', methods=['POST'])
def delete_course(course_id):
    course = Course.query.get_or_404(course_id)
    db.session.delete(course)
    db.session.commit()
    flash('Course deleted successfully!', category='success')
    return redirect(url_for('teacher.courses'))

#Lessons
@teacher.route('/courses/<int:course_id>/lessons', methods=['GET'])
def lessons(course_id):
    teacher_id = current_user.id
    return f"Lessons for course ID: {course_id}"

@teacher.route('/courses/<int:course_id>/add_lesson', methods=['GET', 'POST'])
@login_required
def add_lesson(course_id):
    course = Course.query.filter_by(id=course_id, tutor_id=current_user.id).first_or_404()
    if request.method == 'POST':
        # Получаем данные из формы
        topic = request.form.get('topic')
        lesson_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        start_time = datetime.strptime(request.form.get('start_time'), '%H:%M').time()
        end_time = datetime.strptime(request.form.get('end_time'), '%H:%M').time() if request.form.get(
            'end_time') else None
        location = request.form.get('location')

        # Создаём новый урок
        new_lesson = Lesson(
            course_id=course.id,
            topic=topic,
            date=lesson_date,
            start_time=start_time,
            end_time=end_time,
            location=location
        )
        db.session.add(new_lesson)
        db.session.commit()
        return redirect(url_for('teacher.course_details', course_id=course.id))
    return render_template('teacher/add_lesson.html', course=course)

@teacher.route('/add_lesson_general', methods=['GET', 'POST'])
@login_required
def add_lesson_general():
    courses = Course.query.filter_by(tutor_id=current_user.id).all()
    if request.method == 'POST':
        course_id = request.form.get('course_id')
        topic = request.form.get('topic')
        lesson_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        start_time = datetime.strptime(request.form.get('start_time'), '%H:%M').time()
        end_time = datetime.strptime(request.form.get('end_time'), '%H:%M').time() if request.form.get('end_time') else None
        location = request.form.get('location')

        # Создаём новый урок
        new_lesson = Lesson(
            course_id=course_id,
            topic=topic,
            date=lesson_date,
            start_time=start_time,
            end_time=end_time,
            location=location
        )
        db.session.add(new_lesson)
        db.session.commit()
        return redirect(url_for('teacher.dashboard'))
    return render_template('teacher/add_lesson_general.html', courses=courses)

@teacher.route('/courses/<int:course_id>/students', methods=['GET'])
def view_students(course_id):
    teacher_id = current_user.id
    course = Course.query.filter_by(id=course_id, tutor_id=teacher_id).first()
    if not course:
        flash('Course not found!', category='error')
        return redirect(url_for('teacher.courses'))
    students = course.students
    return render_template('teacher/view_students.html', course=course, students=students)

@teacher.route('/courses/<int:course_id>/feedback', methods=['GET'])
def analyze_feedback(course_id):
    teacher_id = current_user.id  # Заглушка для текущего учителя
    course = Course.query.filter_by(id=course_id, tutor_id=teacher_id).first()
    if not course:
        flash('Course not found!', category='error')
        return redirect(url_for('teacher.courses'))

    feedback = course.feedback  # Получение всех фидбеков курса
    feedback_summary = {
        "average_rating": round(sum(f.mark for f in feedback) / len(feedback), 2) if feedback else 0,
        "total_comments": len([f.comment for f in feedback if f.comment]),
        "positive_comments": len([f for f in feedback if f.mark >= 4]),
        "negative_comments": len([f for f in feedback if f.mark < 4]),
    }

    return render_template('teacher/analyze_feedback.html', course=course, feedback_summary=feedback_summary)

#Attendance
# @teacher.route('/attendance', methods=['GET', 'POST'])
# @login_required
# def attendance():
#     from app.models import User
#     teacher_id = current_user.id
#     # Получение всех курсов, связанных с учителем
#     courses = Course.query.filter_by(tutor_id=teacher_id).all()
#
#     # Получение всех уроков, связанных с курсами этого учителя
#     records = db.session.query(Attendance).join(Lesson).join(Course).join(User).all()
#
#     # Получение фильтров из запроса
#     course_filter = request.args.getlist('course')  # Список курсов
#     date_filter = request.args.getlist('date')  # Список дат
#     status_filter = request.args.getlist('status')  # Статус (присутствие/отсутствие)
#
#     query = Lesson.query
#
#     if course_filter:
#         query = query.join(Course).filter(Course.id.in_(course_filter))
#
#     if date_filter:
#         query = query.filter(Lesson.date.in_(date_filter))
#
#     if status_filter:
#         query = query.join(Attendance).filter(Attendance.status.in_(status_filter))
#
#     lessons = query.all()
#
#     # Возврат в шаблон с фильтрами и уроками
#     return render_template('teacher/attendance.html', courses=courses, lessons=lessons,
#                            course_filter=course_filter, date_filter=date_filter, status_filter=status_filter)
#
# @teacher.route('/attendance/<int:lesson_id>', methods=['GET', 'POST'])
# @login_required
# def manage_attendance(lesson_id):
#     from app.models import User
#     # Получаем урок
#     lesson = Lesson.query.get(lesson_id)
#     if not lesson:
#         flash('Lesson not found!', category='error')
#         return redirect(url_for('teacher.attendance'))
#
#     # Получаем курс, связанный с уроком
#     course = Course.query.get(lesson.course_id)
#     if not course:
#         flash('Course not found!', category='error')
#         return redirect(url_for('teacher.attendance'))
#
#     # Получаем студентов и их посещаемость
#     attendance_records = Attendance.query.filter_by(lesson_id=lesson_id).all()
#
#     if request.method == 'POST':
#         # Обработка изменения статуса студентов
#         for record in attendance_records:
#             status = request.form.get(f'status_{record.id}')
#             comment = request.form.get(f'comment_{record.id}')
#             reason = request.form.get(f'reason_{record.id}')
#             record.status = status
#             record.comments = comment
#             record.reason_of_excuse = reason
#         db.session.commit()
#         flash('Attendance updated successfully!', category='success')
#         return redirect(url_for('teacher.manage_attendance', lesson_id=lesson_id))
#
#     return render_template('teacher/manage_attendance.html', lesson=lesson, course=course,
#                            attendance_records=attendance_records)

@teacher.route('/attendance', methods=['GET'])
def attendance():
    courses = Course.query.all()
    lessons = Lesson.query.all()

    # Чтение параметров фильтра из запроса
    course_filter = request.args.getlist('course')
    date_filter = request.args.getlist('date')  # Теперь это список
    status_filter = request.args.getlist('status')

    # Базовый запрос
    query = Attendance.query

    if course_filter:
        query = query.join(Lesson).filter(Lesson.course_id.in_(course_filter))
    if date_filter:
        query = query.filter(Attendance.lesson.has(Lesson.date.in_(date_filter)))
    if status_filter:
        query = query.filter(Attendance.status.in_(status_filter))

    attendance_records = query.all()

    return render_template(
        'teacher/attendance.html',
        courses=courses,
        lessons=lessons,
        attendance_records=attendance_records,
        course_filter=course_filter,
        date_filter=date_filter,
        status_filter=status_filter,
    )

@teacher.route('/get_lessons_for_course/<int:course_id>', methods=['GET'])
def get_lessons_for_course(course_id):
    lessons = Lesson.query.filter_by(course_id=course_id).all()
    print(f"Lessons for course {course_id}: {lessons}")  # Добавьте для отладки
    lessons_data = [{"date": lesson.date} for lesson in lessons]
    return jsonify({"lessons": lessons_data})

@teacher.route('/attendance/edit/<int:record_id>', methods=['GET', 'POST'])
@login_required
def edit_attendance_record(record_id):
    record = Attendance.query.get_or_404(record_id)

    if request.method == 'POST':
        # Получаем данные из формы
        record.status = request.form.get('status')
        record.comments = request.form.get('comments', '')  # Допустим, есть поле комментариев
        db.session.commit()
        flash('Attendance record updated successfully!', 'success')
        return redirect(url_for('teacher.attendance'))

    return render_template('teacher/edit_attendance.html', record=record)

@teacher.route('/attendance/update_status/<int:record_id>', methods=['POST'])
@login_required
def update_attendance_status(record_id):
    record = Attendance.query.get_or_404(record_id)
    record.status = 'not' if record.status == 'was' else 'was'
    db.session.commit()
    return jsonify({"success": True, "new_status": record.status})

@teacher.route('/update_all_statuses', methods=['POST'])
@login_required
def update_all_statuses():
    data = request.get_json()
    new_status = data.get('status')

    if new_status not in ['was', 'not']:
        return jsonify({'error': 'Invalid status'}), 400

    # Обновляем все записи статусов
    records = Attendance.query.join(Lesson).join(Course).filter(Course.tutor_id == current_user.id).all()
    for record in records:
        if record.status != new_status:  # Меняем статус только если он отличается
            record.status = new_status
    db.session.commit()

    return jsonify({'success': True}), 200

#Feedback
@teacher.route('/feedback', methods=['GET'])
def feedback():
    teacher_id = current_user.id
    courses = Course.query.filter_by(tutor_id=teacher_id).all()

    # Получение параметров фильтрации
    selected_course_id = request.args.get('course_id', type=int)
    specific_date = request.args.get('specific_date')
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'asc')

    # Базовый запрос
    query = Feedback.query.join(Lesson).join(Course).filter(Course.tutor_id == teacher_id)

    # Фильтрация по курсу
    if selected_course_id:
        query = query.filter(Course.id == selected_course_id)

    if specific_date:
        query = query.filter(Lesson.date == specific_date)

    # Сортировка
    if sort_by == 'mark':
        query = query.order_by(Feedback.mark.desc() if sort_order == 'desc' else Feedback.mark.asc())
    else:
        query = query.order_by(Lesson.date.desc() if sort_order == 'desc' else Lesson.date.asc())

    feedback_records = query.all()

    return render_template(
        'teacher/feedback.html',
        courses=courses,
        feedback_records=feedback_records,
        selected_course_id=selected_course_id,
        specific_date=specific_date,
        sort_by=sort_by,
        sort_order=sort_order
    )

@teacher.route('/feedback/response/<int:feedback_id>', methods=['POST'])
def reply_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    response = request.form['response']
    feedback.response_on_feedback = response  # Ответ преподавателя

    db.session.commit()

    flash('Response added successfully!', 'success')
    return redirect(url_for('teacher.feedback'))

@teacher.route('/feedback/hide/<int:feedback_id>', methods=['POST'])
def hide_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    feedback.is_hidden = True  # Скрываем фидбек (нужно добавить поле в модели)

    db.session.commit()

    flash('Feedback hidden successfully!', 'success')
    return redirect(url_for('teacher.feedback'))

@teacher.route('/feedback/unhide/<int:feedback_id>', methods=['POST'])

def unhide_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    feedback.is_hidden = False  # Восстанавливаем фидбек

    db.session.commit()

    flash('Feedback restored successfully!', 'success')
    return redirect(url_for('teacher.feedback'))

@teacher.route('/feedback/<int:lesson_id>', methods=['GET', 'POST'])
def submit_feedback(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    feedbacks = Feedback.query.filter_by(lesson_id=lesson_id).all()
    if request.method == 'POST':
        # Получаем данные формы
        mark = request.form['mark']
        comment = request.form['comment']
        anonymous = 'anonymous' in request.form
        feedback_type = request.form['type']
        student_id = request.form['student_id']  # Здесь получаем ID студента из формы

        # Создаем новый объект Feedback
        feedback = Feedback(
            lesson_id=lesson.id,
            student_id=student_id,  # Используем переданный ID студента
            mark=mark,
            comment=comment,
            anonymous=anonymous,
            type=feedback_type
        )

        db.session.add(feedback)
        db.session.commit()

        flash('Feedback submitted successfully!', 'success')
        return redirect(url_for('teacher.feedback', lesson_id=lesson.id))

    return render_template('teacher/manage_feedback.html', lesson=lesson)

@teacher.route('/feedback/delete/<int:feedback_id>', methods=['POST'])
def delete_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    db.session.delete(feedback)
    db.session.commit()

    flash('Feedback deleted successfully!', 'success')
    return redirect(url_for('teacher.feedback'))

#Export

@teacher.route('/export/csv', methods=['GET'])
def export_csv():
    from app.models import User
    teacher_id = current_user.id

    # Получаем параметры фильтрации
    course_filter = request.args.get('course', None)
    start_date_str = request.args.get('start_date', None)
    end_date_str = request.args.get('end_date', None)
    student_filter = request.args.get('student', None)

    # Преобразование дат
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

    # Запрос для фильтрации Feedback
    query = Feedback.query.join(Course).join(User, Feedback.student_id == User.id).filter(Course.tutor_id == teacher_id)

    if course_filter:
        query = query.filter(Course.id == course_filter)
    if start_date:
        query = query.filter(Feedback.exact_time >= start_date)
    if end_date:
        query = query.filter(Feedback.exact_time <= end_date)
    if student_filter:
        query = query.filter(User.id == student_filter)

    filtered_feedbacks = query.all()

    # Если записей нет, вернуть пустой CSV с заголовками
    if not filtered_feedbacks:
        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(['Course', 'Student', 'Mark', 'Comment'])
        response = Response(si.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = 'attachment; filename=feedback_export.csv'
        return response

    # Подготовка данных для CSV
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Course', 'Student', 'Mark', 'Comment'])  # Заголовки CSV

    for feedback in filtered_feedbacks:
        writer.writerow([
            feedback.course.name,
            feedback.student.full_name,
            feedback.mark,
            feedback.comment
        ])

    # Создание ответа
    output = si.getvalue()
    response = Response(output, mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=feedback_export.csv'
    return response


@teacher.route('/get_groups', methods=['GET'])
def get_groups():
    from app.models import User
    course_id = request.args.get('course_id')
    # Фильтруем пользователей-студентов по курсу
    groups = User.query.filter_by(role='student').with_entities(User.group).distinct().all()
    group_list = [g.group for g in groups if g.group]
    return jsonify(group_list)

@teacher.route('/get_students', methods=['GET'])
def get_students():
    from app.models import User
    group_name = request.args.get('group_name')  # Группа
    if group_name:
        students = User.query.filter_by(role='student', group=group_name).all()
    else:
        students = User.query.filter_by(role='student').all()

    student_list = [{'id': s.id, 'full_name': s.full_name} for s in students]
    return jsonify(student_list)

@teacher.route('/export', methods=['GET', 'POST'])
def export():
    if request.method == 'POST':
        # Получаем параметры фильтрации из формы
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        data_type = request.form.get('data_type', 'attendance')  # По умолчанию экспорт посещаемости
        course_filter = request.form.get('course')
        group_filter = request.form.get('group')  # Новый фильтр группы
        student_filter = request.form.get('student')  # Новый фильтр студента

        # Преобразуем start_date и end_date в формат datetime.date
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Получаем курсы, которые доступны для учителя
        teacher_id = current_user.id
        courses = Course.query.filter_by(tutor_id=teacher_id).all()

        if data_type == 'attendance':
            attendance_records = []
            for course in courses:
                if course_filter and str(course.id) != course_filter:
                    continue  # Пропустить курс, если фильтр не совпадает
                lessons = Lesson.query.filter_by(course_id=course.id).all()
                for lesson in lessons:
                    attendance_records.extend(Attendance.query.filter_by(lesson_id=lesson.id).all())

            # Фильтрация по датам
            if start_date:
                attendance_records = [a for a in attendance_records if a.lesson.date >= start_date]
            if end_date:
                attendance_records = [a for a in attendance_records if a.lesson.date <= end_date]

            # Фильтрация по группе
            if group_filter and group_filter.lower() != 'все':
                attendance_records = [
                    a for a in attendance_records if group_filter.lower() in (a.student.group or '').lower()
                ]

            # Фильтрация по студенту
            if student_filter and student_filter.lower() != 'все':
                attendance_records = [
                    a for a in attendance_records if student_filter.lower() in (a.student.full_name or '').lower()
                ]

            # Если нет записей для экспорта
            if not attendance_records:
                flash('No records found for the selected filters.', 'warning')
                return redirect(url_for('teacher.export'))

            # Экспорт в Excel
            wb = Workbook()
            ws = wb.active
            ws.append(["Student Name", "Course Name", "Lesson Date", "Status", "Comments", "Reason of Excuse"])

            for record in attendance_records:
                student_name = record.student.full_name
                course_name = record.lesson.course.name
                lesson_date = record.lesson.date
                status = record.status
                comments = record.comments
                reason_of_excuse = record.reason_of_excuse
                ws.append([student_name, course_name, lesson_date, status, comments, reason_of_excuse])

            # Сохранение в буфер
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Отправка файла пользователю
            return send_file(output, as_attachment=True, download_name="svg_attendance_export.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif data_type == 'feedback':
            feedback_records = Feedback.query.join(Course).filter(Course.tutor_id == teacher_id)

            # Фильтрация по курсу
            if course_filter and course_filter.lower() != 'все':
                feedback_records = feedback_records.filter(Feedback.course_id == course_filter)

            # Фильтрация по дате
            if start_date:
                feedback_records = feedback_records.filter(Feedback.exact_time >= start_date)
            if end_date:
                feedback_records = feedback_records.filter(Feedback.exact_time <= end_date)

            feedback_records = feedback_records.all()

            # Если нет записей для экспорта
            if not feedback_records:
                flash('No feedback records found for the selected filters.', 'warning')
                return redirect(url_for('teacher.export'))

            # Экспорт фидбэков в Excel
            wb = Workbook()
            ws = wb.active
            ws.append(["Course Name", "Student Name", "Feedback Mark", "Comment", "Exact Time"])

            for feedback in feedback_records:
                ws.append([
                    feedback.course.name,
                    feedback.student.full_name,
                    feedback.mark,
                    feedback.comment,
                    feedback.exact_time.strftime('%Y-%m-%d %H:%M:%S')
                ])

            # Сохранение в буфер
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Отправка файла пользователю
            return send_file(output, as_attachment=True, download_name="svg_feedback_export.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif data_type == 'course':
            # Экспорт курсов
            course_records = courses if courses else []

            # Если нет записей для экспорта
            if not course_records:
                flash('No course records found for the selected filters.', 'warning')
                return redirect(url_for('teacher.export'))

            # Экспорт курсов в Excel
            wb = Workbook()
            ws = wb.active
            ws.append(["Course Name", "Description", "Start Date", "End Date", "Category", "Number of Students"])

            for course in course_records:
                ws.append([
                    course.name,
                    course.description,
                    course.start_date.strftime('%Y-%m-%d') if course.start_date else "N/A",
                    course.end_date.strftime('%Y-%m-%d') if course.end_date else "N/A",
                    course.category if course.category else "N/A",
                    course.students_count
                ])

            # Сохранение в буфер
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Отправка файла пользователю
            return send_file(output, as_attachment=True, download_name="svg_course_export.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        else:
            flash('Invalid data type selected for export.', 'danger')
            return redirect(url_for('teacher.export'))

    # Для GET-запроса показываем форму с выбором
    teacher_id = current_user.id
    courses = Course.query.filter_by(tutor_id=teacher_id).all()
    return render_template('teacher/export.html', courses=courses)

'''JSONIFY'''

'''
from flask import render_template, request, redirect, url_for, flash, Blueprint, Response, jsonify
from app.models import Course
from app import db
from app.models import Lesson, Attendance, Feedback, Course, Lesson
from datetime import datetime, timedelta
from sqlalchemy.sql import func
#for export
import csv
from io import StringIO
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename
import os
#maintain my man

teacher = Blueprint('teacher', __name__, template_folder='templates')

@teacher.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    from app.models import User
    from app.models import CourseStudent

    current_user_data = User.query.get(current_user.id)
    user_name = current_user_data.full_name
    user_role = current_user_data.role
    total_users = User.query.count()
    total_students = User.query.filter_by(role='student').count()
    total_teachers = User.query.filter_by(role='teacher').count()
    active_users = User.query.filter(User.last_login >= datetime.now() - timedelta(days=1)).count()
    active_courses = Course.query.filter_by(status='active').count()

    # Самый популярный курс
    popular_course_query = db.session.query(
        Course,
        db.func.count(CourseStudent.c.student_id).label('student_count')
    ).join(CourseStudent, Course.id == CourseStudent.c.course_id).group_by(Course.id).order_by(
        db.func.count(CourseStudent.c.student_id).desc()
    ).first()
    popular_course = popular_course_query[0].name if popular_course_query else "No courses yet"

    # Средняя оценка уроков
    avg_lesson_rating = Feedback.query.with_entities(func.avg(Feedback.mark)).scalar()
    teacher_feedback_query = Feedback.query.join(Course, Feedback.course_id == Course.id).filter(
        Course.tutor_id == current_user.id
    ).with_entities(func.avg(Feedback.mark)).scalar()
    teacher_rating = round(teacher_feedback_query, 2) if teacher_feedback_query else "N/A"

    return jsonify({
        'user_name': user_name,
        'user_role': user_role,
        'total_users': total_users,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'active_users': active_users,
        'active_courses': active_courses,
        'popular_course': popular_course,
        'avg_lesson_rating': round(avg_lesson_rating, 2) if avg_lesson_rating else "N/A",
        'teacher_rating': teacher_rating
    }), 200

@teacher.route('/help', methods=['GET'])
@login_required
def help_page():
    return jsonify({'message': 'Help page content.'}), 200

@teacher.route('/tips', methods=['GET'])
@login_required
def tips_page():
    return jsonify({'message': 'Tips page content.'}), 200

@teacher.route('/notifications', methods=['GET'])
@login_required
def notifications():
    return jsonify({'message': 'Notifications content.'}), 200

@teacher.route('/teacher/profile', methods=['GET', 'POST'])
@login_required
def profile():
    from app.models import User
    user = User.query.filter_by(id=current_user.id).first()

    if request.method == 'POST':
        # Собираем данные из формы
        user.full_name = request.form.get('full_name')
        user.nickname = request.form.get('nickname')
        user.university = request.form.get('university')
        user.institute = request.form.get('institute')
        user.phone_number = request.form.get('phone_number')
        user.date_of_birth = request.form.get('date_of_birth') or None

        # Обработка загруженного фото
        if 'profile_picture' in request.files:
            picture = request.files['profile_picture']
            if picture:
                picture_filename = secure_filename(picture.filename)
                picture_path = os.path.join('static', 'images', picture_filename)
                picture.save(picture_path)
                user.profile_picture = picture_path

        db.session.commit()
        return jsonify({'message': 'Profile updated successfully!'}), 200

    # Возвращаем данные профиля в формате JSON
    return jsonify({
        'full_name': user.full_name,
        'nickname': user.nickname,
        'university': user.university,
        'institute': user.institute,
        'phone_number': user.phone_number,
        'date_of_birth': user.date_of_birth,
        'profile_picture': user.profile_picture
    }), 200

@teacher.route('/teacher/update_profile', methods=['POST'])
@login_required
def update_profile():
    from app.models import User
    user = User.query.filter_by(id=current_user.id).first()

    if not user:
        return jsonify({'error': 'User not found'}), 404

    data = request.get_json()

    # Обновление данных пользователя
    user.full_name = data.get('full_name', user.full_name)
    user.nickname = data.get('nickname', user.nickname)
    user.university = data.get('university', user.university)
    user.institute = data.get('institute', user.institute)
    user.phone_number = data.get('phone_number', user.phone_number)
    user.date_of_birth = data.get('date_of_birth', user.date_of_birth)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Profile updated successfully',
        'updated_user': {
            'full_name': user.full_name,
            'nickname': user.nickname,
            'university': user.university,
            'institute': user.institute,
            'phone_number': user.phone_number,
            'date_of_birth': user.date_of_birth
        }
    }), 200


@teacher.route('/courses', methods=['GET'])
@login_required
def courses():
    teacher_id = current_user.id
    courses = Course.query.filter_by(tutor_id=teacher_id).all()
    courses_data = [{
        'id': course.id,
        'name': course.name,
        'description': course.description,
        'start_date': course.start_date.strftime('%Y-%m-%d') if course.start_date else None,
        'end_date': course.end_date.strftime('%Y-%m-%d') if course.end_date else None,
        'course_code': course.course_code,
        'category': course.category,
        'students_count': course.students_count
    } for course in courses]

    return jsonify({'courses': courses_data})

@teacher.route('/courses/add', methods=['POST'])
@login_required
def add_course():
    data = request.get_json()
    name = data.get('name')
    description = data.get('description')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    course_code = data.get('course_code')
    category = data.get('category')

    teacher_id = current_user.id  # ID текущего преподавателя

    new_course = Course(
        name=name,
        description=description,
        tutor_id=teacher_id,
        start_date=start_date,
        end_date=end_date,
        course_code=course_code,
        category=category
    )
    db.session.add(new_course)
    db.session.commit()

    return jsonify({'message': 'Course added successfully!', 'course_id': new_course.id})


@teacher.route('/courses/<int:course_id>', methods=['GET'])
@login_required
def course_details(course_id):
    course = Course.query.filter_by(id=course_id, tutor_id=current_user.id).first_or_404()
    lessons = Lesson.query.filter_by(course_id=course_id).order_by(Lesson.date).all()

    total_lessons = len(lessons)
    completed_lessons = sum(1 for lesson in lessons if lesson.date <= datetime.utcnow().date())
    remaining_lessons = total_lessons - completed_lessons

    course_data = {
        'id': course.id,
        'name': course.name,
        'description': course.description,
        'start_date': course.start_date.strftime('%Y-%m-%d') if course.start_date else None,
        'end_date': course.end_date.strftime('%Y-%m-%d') if course.end_date else None,
        'course_code': course.course_code,
        'category': course.category
    }
    lessons_data = [{
        'id': lesson.id,
        'date': lesson.date.strftime('%Y-%m-%d'),
        'topic': lesson.topic
    } for lesson in lessons]

    return jsonify({
        'course': course_data,
        'lessons': lessons_data,
        'statistics': {
            'total_lessons': total_lessons,
            'completed_lessons': completed_lessons,
            'remaining_lessons': remaining_lessons
        }
    })


@teacher.route('/courses/<int:course_id>/edit', methods=['POST'])
@login_required
def edit_course(course_id):
    data = request.get_json()
    course = Course.query.get_or_404(course_id)

    course.name = data.get('name')
    course.description = data.get('description')
    course.start_date = data.get('start_date')
    course.end_date = data.get('end_date')
    course.course_code = data.get('course_code')
    course.category = data.get('category')

    db.session.commit()

    return jsonify({'message': 'Course updated successfully!'})


@teacher.route('/courses/<int:course_id>/delete', methods=['DELETE'])
@login_required
def delete_course(course_id):
    course = Course.query.get_or_404(course_id)
    db.session.delete(course)
    db.session.commit()

    return jsonify({'message': 'Course deleted successfully!'})

@teacher.route('/courses/<int:course_id>/lessons', methods=['GET'])
@login_required
def lessons(course_id):
    teacher_id = current_user.id
    lessons = Lesson.query.filter_by(course_id=course_id).all()
    lessons_data = [{
        'id': lesson.id,
        'topic': lesson.topic,
        'date': lesson.date.strftime('%Y-%m-%d'),
        'start_time': lesson.start_time.strftime('%H:%M'),
        'end_time': lesson.end_time.strftime('%H:%M') if lesson.end_time else None,
        'location': lesson.location
    } for lesson in lessons]

    return jsonify({'lessons': lessons_data})


@teacher.route('/courses/<int:course_id>/add_lesson', methods=['POST'])
@login_required
def add_lesson(course_id):
    data = request.get_json()
    topic = data.get('topic')
    lesson_date = datetime.strptime(data.get('date'), '%Y-%m-%d').date()
    start_time = datetime.strptime(data.get('start_time'), '%H:%M').time()
    end_time = datetime.strptime(data.get('end_time'), '%H:%M').time() if data.get('end_time') else None
    location = data.get('location')

    new_lesson = Lesson(
        course_id=course_id,
        topic=topic,
        date=lesson_date,
        start_time=start_time,
        end_time=end_time,
        location=location
    )
    db.session.add(new_lesson)
    db.session.commit()

    return jsonify({'message': 'Lesson added successfully!', 'lesson_id': new_lesson.id})


@teacher.route('/add_lesson_general', methods=['POST'])
@login_required
def add_lesson_general():
    data = request.get_json()
    course_id = data.get('course_id')
    topic = data.get('topic')
    lesson_date = datetime.strptime(data.get('date'), '%Y-%m-%d').date()
    start_time = datetime.strptime(data.get('start_time'), '%H:%M').time()
    end_time = datetime.strptime(data.get('end_time'), '%H:%M').time() if data.get('end_time') else None
    location = data.get('location')

    new_lesson = Lesson(
        course_id=course_id,
        topic=topic,
        date=lesson_date,
        start_time=start_time,
        end_time=end_time,
        location=location
    )
    db.session.add(new_lesson)
    db.session.commit()

    return jsonify({'message': 'Lesson added successfully!', 'lesson_id': new_lesson.id})


@teacher.route('/courses/<int:course_id>/students', methods=['GET'])
@login_required
def view_students(course_id):
    teacher_id = current_user.id
    course = Course.query.filter_by(id=course_id, tutor_id=teacher_id).first_or_404()
    students_data = [{
        'id': student.id,
        'full_name': student.full_name,
        'email': student.email
    } for student in course.students]

    return jsonify({'course_id': course.id, 'students': students_data})


@teacher.route('/courses/<int:course_id>/feedback', methods=['GET'])
@login_required
def analyze_feedback(course_id):
    teacher_id = current_user.id
    course = Course.query.filter_by(id=course_id, tutor_id=teacher_id).first_or_404()
    feedback = course.feedback

    feedback_summary = {
        "average_rating": round(sum(f.mark for f in feedback) / len(feedback), 2) if feedback else 0,
        "total_comments": len([f.comment for f in feedback if f.comment]),
        "positive_comments": len([f for f in feedback if f.mark >= 4]),
        "negative_comments": len([f for f in feedback if f.mark < 4])
    }

    feedback_data = [{
        'id': f.id,
        'mark': f.mark,
        'comment': f.comment,
        'student_id': f.student_id,
        'lesson_id': f.lesson_id
    } for f in feedback]

    return jsonify({'course_id': course.id, 'feedback_summary': feedback_summary, 'feedback': feedback_data})


@teacher.route('/attendance', methods=['GET'])
@login_required
def attendance():
    # Получаем фильтры из запроса
    course_filter = request.args.getlist('course')
    date_filter = request.args.getlist('date')
    status_filter = request.args.getlist('status')

    # Базовый запрос для посещаемости
    query = Attendance.query

    if course_filter and course_filter != ['']:
        query = query.join(Lesson).filter(Lesson.course_id.in_(course_filter))
    if date_filter and date_filter != ['']:
        query = query.filter(Attendance.lesson.has(Lesson.date.in_(date_filter)))
    if status_filter and status_filter != ['']:
        query = query.filter(Attendance.status.in_(status_filter))

    # Получаем записи посещаемости
    attendance_records = query.all()
    attendance_data = [{
        'id': record.id,
        'lesson_id': record.lesson_id,
        'student_id': record.student_id,
        'status': record.status,
        'comments': record.comments,
        'lesson_date': record.lesson.date.strftime('%Y-%m-%d') if record.lesson else None
    } for record in attendance_records]

    return jsonify({
        'attendance_records': attendance_data,
        'filters': {
            'course': course_filter,
            'date': date_filter,
            'status': status_filter
        }
    })


@teacher.route('/get_lessons_for_course/<int:course_id>', methods=['GET'])
@login_required
def get_lessons_for_course(course_id):
    lessons = Lesson.query.filter_by(course_id=course_id).all()
    lessons_data = [{"id": lesson.id, "date": lesson.date.strftime('%Y-%m-%d')} for lesson in lessons]

    return jsonify({"lessons": lessons_data})


@teacher.route('/attendance/edit/<int:record_id>', methods=['POST'])
@login_required
def edit_attendance_record(record_id):
    record = Attendance.query.get_or_404(record_id)
    data = request.get_json()

    record.status = data.get('status', record.status)
    record.comments = data.get('comments', record.comments)
    db.session.commit()

    return jsonify({"message": "Attendance record updated successfully!"})


@teacher.route('/attendance/update_status/<int:record_id>', methods=['POST'])
@login_required
def update_attendance_status(record_id):
    record = Attendance.query.get_or_404(record_id)
    record.status = 'not' if record.status == 'was' else 'was'
    db.session.commit()

    return jsonify({"success": True, "new_status": record.status})


@teacher.route('/update_all_statuses', methods=['POST'])
@login_required
def update_all_statuses():
    data = request.get_json()
    new_status = data.get('status')

    if new_status not in ['was', 'not']:
        return jsonify({'error': 'Invalid status'}), 400

    # Обновляем все записи статусов
    records = Attendance.query.join(Lesson).join(Course).filter(Course.tutor_id == current_user.id).all()
    for record in records:
        if record.status != new_status:  # Меняем статус только если он отличается
            record.status = new_status
    db.session.commit()

    return jsonify({'success': True})


@teacher.route('/feedback', methods=['GET'])
@login_required
def feedback():
    teacher_id = current_user.id
    courses = Course.query.filter_by(tutor_id=teacher_id).all()

    # Получение параметров фильтрации
    selected_course_id = request.args.get('course_id', type=int)
    specific_date = request.args.get('specific_date')
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'asc')

    # Базовый запрос
    query = Feedback.query.join(Lesson).join(Course).filter(Course.tutor_id == teacher_id)

    # Фильтрация по курсу
    if selected_course_id:
        query = query.filter(Course.id == selected_course_id)

    if specific_date:
        query = query.filter(Lesson.date == specific_date)

    # Сортировка
    if sort_by == 'mark':
        query = query.order_by(Feedback.mark.desc() if sort_order == 'desc' else Feedback.mark.asc())
    else:
        query = query.order_by(Lesson.date.desc() if sort_order == 'desc' else Lesson.date.asc())

    feedback_records = query.all()

    feedback_data = [{
        'id': feedback.id,
        'lesson_id': feedback.lesson_id,
        'student_id': feedback.student_id,
        'mark': feedback.mark,
        'comment': feedback.comment,
        'anonymous': feedback.anonymous,
        'type': feedback.type,
        'is_hidden': feedback.is_hidden,
        'response': feedback.response_on_feedback,
        'lesson_date': feedback.lesson.date.strftime('%Y-%m-%d') if feedback.lesson.date else None,
        'course_name': feedback.lesson.course.name if feedback.lesson.course else None
    } for feedback in feedback_records]

    courses_data = [{'id': course.id, 'name': course.name} for course in courses]

    return jsonify({
        'feedback_records': feedback_data,
        'courses': courses_data,
        'filters': {
            'selected_course_id': selected_course_id,
            'specific_date': specific_date,
            'sort_by': sort_by,
            'sort_order': sort_order
        }
    })


@teacher.route('/feedback/response/<int:feedback_id>', methods=['POST'])
@login_required
def reply_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    data = request.get_json()
    response = data.get('response')
    feedback.response_on_feedback = response

    db.session.commit()

    return jsonify({'success': True, 'message': 'Response added successfully!'})


@teacher.route('/feedback/hide/<int:feedback_id>', methods=['POST'])
@login_required
def hide_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    feedback.is_hidden = True

    db.session.commit()

    return jsonify({'success': True, 'message': 'Feedback hidden successfully!'})


@teacher.route('/feedback/unhide/<int:feedback_id>', methods=['POST'])
@login_required
def unhide_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    feedback.is_hidden = False

    db.session.commit()

    return jsonify({'success': True, 'message': 'Feedback restored successfully!'})


@teacher.route('/feedback/<int:lesson_id>', methods=['POST'])
@login_required
def submit_feedback(lesson_id):
    data = request.get_json()
    lesson = Lesson.query.get_or_404(lesson_id)

    mark = data.get('mark')
    comment = data.get('comment')
    anonymous = data.get('anonymous', False)
    feedback_type = data.get('type')
    student_id = data.get('student_id')

    feedback = Feedback(
        lesson_id=lesson.id,
        student_id=student_id,
        mark=mark,
        comment=comment,
        anonymous=anonymous,
        type=feedback_type
    )

    db.session.add(feedback)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Feedback submitted successfully!'})


@teacher.route('/feedback/delete/<int:feedback_id>', methods=['POST'])
@login_required
def delete_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    db.session.delete(feedback)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Feedback deleted successfully!'})


@teacher.route('/export/csv', methods=['GET'])
@login_required
def export_csv():
    from app.models import User
    teacher_id = current_user.id

    # Получаем параметры фильтрации
    course_filter = request.args.get('course', None)
    start_date_str = request.args.get('start_date', None)
    end_date_str = request.args.get('end_date', None)
    student_filter = request.args.get('student', None)

    # Преобразование дат
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

    # Запрос для фильтрации Feedback
    query = Feedback.query.join(Course).join(User, Feedback.student_id == User.id).filter(Course.tutor_id == teacher_id)

    if course_filter:
        query = query.filter(Course.id == course_filter)
    if start_date:
        query = query.filter(Feedback.exact_time >= start_date)
    if end_date:
        query = query.filter(Feedback.exact_time <= end_date)
    if student_filter:
        query = query.filter(User.id == student_filter)

    filtered_feedbacks = query.all()

    # Если записей нет, вернуть JSON с сообщением
    if not filtered_feedbacks:
        return jsonify({'error': 'No feedbacks found for the given filters'}), 404

    # Формат данных для возврата JSON
    feedback_data = [
        {
            'course': feedback.course.name,
            'student': feedback.student.full_name,
            'mark': feedback.mark,
            'comment': feedback.comment
        }
        for feedback in filtered_feedbacks
    ]

    if 'format' in request.args and request.args['format'] == 'json':
        return jsonify(feedback_data)

    # Подготовка данных для CSV
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Course', 'Student', 'Mark', 'Comment'])  # Заголовки CSV

    for feedback in filtered_feedbacks:
        writer.writerow([
            feedback.course.name,
            feedback.student.full_name,
            feedback.mark,
            feedback.comment
        ])

    # Создание ответа
    output = si.getvalue()
    response = Response(output, mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=feedback_export.csv'
    return response


@teacher.route('/get_groups', methods=['GET'])
@login_required
def get_groups():
    from app.models import User
    try:
        course_id = request.args.get('course_id')
        groups = User.query.filter_by(role='student').with_entities(User.group).distinct().all()
        if not groups:
            return jsonify({'error': 'No groups found'}), 404

        group_list = [g.group for g in groups if g.group]
        return jsonify(group_list)
    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500


@teacher.route('/get_students', methods=['GET'])
@login_required
def get_students():
    from app.models import User
    try:
        group_name = request.args.get('group_name')  # Группа
        if group_name:
            students = User.query.filter_by(role='student', group=group_name).all()
        else:
            students = User.query.filter_by(role='student').all()

        if not students:
            return jsonify({'error': 'No students found'}), 404

        student_list = [{'id': s.id, 'full_name': s.full_name} for s in students]
        return jsonify(student_list)
    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
#main export

@teacher.route('/export', methods=['GET', 'POST'])
@login_required
def export():
    try:
        if request.method == 'POST':
            # Получаем параметры фильтрации из формы
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            data_type = request.form.get('data_type', 'attendance')  # По умолчанию экспорт посещаемости
            course_filter = request.form.get('course')
            group_filter = request.form.get('group')
            student_filter = request.form.get('student')

            # Преобразуем start_date и end_date в формат datetime.date
            start_date = None
            end_date = None
            if start_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return jsonify({'error': 'Invalid start date format. Use YYYY-MM-DD.'}), 400
            if end_date_str:
                try:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return jsonify({'error': 'Invalid end date format. Use YYYY-MM-DD.'}), 400

            # Получаем курсы, которые доступны для учителя
            teacher_id = current_user.id
            courses = Course.query.filter_by(tutor_id=teacher_id).all()

            if data_type == 'attendance':
                attendance_records = []
                for course in courses:
                    if course_filter and str(course.id) != course_filter:
                        continue
                    lessons = Lesson.query.filter_by(course_id=course.id).all()
                    for lesson in lessons:
                        attendance_records.extend(Attendance.query.filter_by(lesson_id=lesson.id).all())

                # Фильтрация по датам
                if start_date:
                    attendance_records = [a for a in attendance_records if a.lesson.date >= start_date]
                if end_date:
                    attendance_records = [a for a in attendance_records if a.lesson.date <= end_date]

                # Фильтрация по группе
                if group_filter and group_filter.lower() != 'все':
                    attendance_records = [
                        a for a in attendance_records if group_filter.lower() in (a.student.group or '').lower()
                    ]

                # Фильтрация по студенту
                if student_filter and student_filter.lower() != 'все':
                    attendance_records = [
                        a for a in attendance_records if student_filter.lower() in (a.student.full_name or '').lower()
                    ]

                if not attendance_records:
                    return jsonify({'error': 'No attendance records found for the selected filters.'}), 404

                attendance_data = [
                    {
                        'student_name': record.student.full_name,
                        'course_name': record.lesson.course.name,
                        'lesson_date': record.lesson.date.isoformat(),
                        'status': record.status,
                        'comments': record.comments,
                        'reason_of_excuse': record.reason_of_excuse
                    }
                    for record in attendance_records
                ]

                return jsonify({'attendance_records': attendance_data})

            elif data_type == 'feedback':
                feedback_records = Feedback.query.join(Course).filter(Course.tutor_id == teacher_id)

                if course_filter and course_filter.lower() != 'все':
                    feedback_records = feedback_records.filter(Feedback.course_id == course_filter)

                if start_date:
                    feedback_records = feedback_records.filter(Feedback.exact_time >= start_date)
                if end_date:
                    feedback_records = feedback_records.filter(Feedback.exact_time <= end_date)

                feedback_records = feedback_records.all()

                if not feedback_records:
                    return jsonify({'error': 'No feedback records found for the selected filters.'}), 404

                feedback_data = [
                    {
                        'course_name': feedback.course.name,
                        'student_name': feedback.student.full_name,
                        'mark': feedback.mark,
                        'comment': feedback.comment,
                        'exact_time': feedback.exact_time.strftime('%Y-%m-%d %H:%M:%S')
                    }
                    for feedback in feedback_records
                ]

                return jsonify({'feedback_records': feedback_data})

            elif data_type == 'course':
                if not courses:
                    return jsonify({'error': 'No courses found for the teacher.'}), 404

                course_data = [
                    {
                        'course_name': course.name,
                        'description': course.description,
                        'start_date': course.start_date.isoformat() if course.start_date else None,
                        'end_date': course.end_date.isoformat() if course.end_date else None,
                        'category': course.category,
                        'students_count': course.students_count
                    }
                    for course in courses
                ]

                return jsonify({'courses': course_data})

            else:
                return jsonify({'error': 'Invalid data type selected for export.'}), 400

        # Для GET-запроса показываем курсы
        teacher_id = current_user.id
        courses = Course.query.filter_by(tutor_id=teacher_id).all()
        course_data = [{'id': course.id, 'name': course.name} for course in courses]
        return jsonify({'courses': course_data})

    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
'''

'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''
'''---------------------------------------------------------------------'''

import json
from flask import Blueprint, request, redirect,render_template
from app.models import Course, User, CourseStudent, Lesson, Attendance, Feedback
from datetime import date
from app import db

@teacher.route('/student/dashboard/<int:id>', methods=['GET'])
def student(id):
    student=User.query.get(id)
    attended_courses=CourseStudent.query.filter_by(student_id=id).all()
    all_lessons=[]
    for i in attended_courses:
        course_lessons=Lesson.query.filter_by(course_id=i.course_id).all()
        all_lessons+=course_lessons
    all_lessons=sorted(all_lessons,key=lambda h: h.date)
    notifications_of_nearest_lessons={}
    count=0
    for i in all_lessons:
        if (0<=(i.date-date.today()).days<=7):
            notifications_of_nearest_lessons.update({f'{count}':{
                'name_of_lesson':str(i.topic),
                'date_of_lesson':str(i.date),
                'start_time':str(i.start_time),
                'location':str(i.location)
            }})
            count+=1
    notifications_of_feedback = {}
    count = 0
    for i in all_lessons:
        if (0 < (date.today()-i.date).days <= 7):
            notifications_of_feedback.update({f'{count}': {
                'name_of_lesson': str(i.topic),
                'date_of_lesson': str(i.date),
            }})
            count += 1
    return {'notifications_of_nearest_lessons':notifications_of_nearest_lessons,
            'general_inf':{'name': student.full_name,'role':student.role},
            'notifications_of_feedback':notifications_of_feedback
            }

@teacher.route('/student/<int:id>/feedback',methods=['POST','GET'])
def student_feedback(id):
    attendance = Attendance.query.filter_by(student_id=id).all()
    if request.method=='GET':
        inf={}
        for i in range(len(attendance)):
            feedback={'lesson_id':str(attendance[i].lesson_id),
                          'course_name':str(attendance[i].lesson.course.name),
                                   'tutor_fullname':str(User.query.filter_by(id=attendance[
                                       i].lesson.course.tutor_id).first().full_name),
                                    'date':str(attendance[i].lesson.date)
                                    }
            if not(Feedback.query.filter_by(student_id=id,lesson_id=attendance[i].lesson.id).first())==None:
                add_feedback={'mark':str(Feedback.query.filter_by(student_id=id,lesson_id=attendance[i].lesson.id).first().mark),
                          'comment':str(Feedback.query.filter_by(student_id=id,lesson_id=attendance[i].lesson.id).first().comment)
                          }
            else:
                add_feedback={'mark':'','comment':''}
            feedback.update(add_feedback)
            inf.update({f'{i}':feedback})
        sorted_inf=dict(
            sorted(inf.items(), key=lambda item: item[1]['date']) #кортеж пары, сортировка по ключу второй элемент
            # кортежа ключ-date
            )
        return sorted_inf
    else:
        mark=request.form['mark']
        comment=request.form['comment']
        lesson_id = request.form['lesson_id'] #как нибудь это передать скрыто
        new_feedback=Feedback(lesson_id=lesson_id,student_id=id,
                              course_id=Lesson.query.filter_by(id=lesson_id).first().course_id,
                              mark=mark, comment=comment)
        db.session.add(new_feedback)
        db.session.commit()
        return redirect(f'/student/{id}/feedback')

@teacher.route('/student/<int:id>/attendance')
def student_attendance(id):
    attendance=Attendance.query.filter_by(student_id=id).all()
    attendance_inf={}
    for i in range(len(attendance)):
        tutor=User.query.filter_by(id=attendance[i].lesson.course.tutor_id).first()
        attendance_inf.update({f'entry_{i}':{
            'tutor_name': str(tutor.full_name),
            'course':str(attendance[i].lesson.course.name),
            'date':str(attendance[i].lesson.date),
            'status':str(attendance[i].status)
                                         }})

    return attendance_inf

@teacher.route('/student/<int:id>/courses')
def student_courses(id):
    student=User.query.get(id)
    att_c=CourseStudent.query.filter_by(student_id=student.id).all()
    att_courses={}
    for i in range(len(att_c)):
        att_courses.update({f"att_course_{i}":{"id": str(att_c[i].course_id),
                                       "name":str(att_c[i].course.name),
                                       'count': str(len(CourseStudent.query.filter_by(course_id=att_c[i].course_id).all())),
                                       'start_date':str(att_c[i].course.start_date),
                                       'end_date': str(att_c[i].course.end_date),
                                       "status": str(att_c[i].course.status),
                                       }})

    att_c_id=[cour.course_id for cour in att_c] #id всех посещаемых курсов
    aff_c_id=[cour.id for cour in Course.query.all() if cour.id not in att_c_id]
    #id всех доступных курсов
    aff_courses={}
    for i in range(len(aff_c_id)):
        course=Course.query.filter_by(id=aff_c_id[i]).first()
        aff_courses.update({f"aff_course_{i}":{"id": str(course.id),
                                       "name":str(course.name),
                                       'count': str(len(CourseStudent.query.filter_by(course_id=course.id).all())),
                                       'start_date':str(course.start_date),
                                       'end_date': str(course.end_date),
                                       "status": str(course.status),
                                       }})
    return {'att_courses':att_courses,'aff_courses':aff_courses}

#кнопка записаться на курс будет в подробнее каждого курса
@teacher.route('/student/<int:id>/courses/<int:course_id>',methods=['GET','POST'])
def student_all_inf_about_course(id,course_id):
    if request.method=='POST':
        new_course_entry=CourseStudent(course_id=course_id,student_id=id)
        try:
            db.session.add(new_course_entry)
            db.session.commit()
            return 'Вы записались на курс'
        except:
            return 'При записи на курс произошла ошибка'
    else:
        course=Course.query.filter_by(id=course_id).first()
        current_date=date.today()
        lessons=Lesson.query.filter_by(course_id=course_id).all()
        count=len([1 for i in lessons if i.date<current_date])
        inf={'course_id':course.id,
             'course_name':course.name,
             'amount_of_finish_lessons': str(count),
             'amount_of_lessons': str(len(lessons)),
             'amount_of_finish_lessons_in_perсentages': str(count/len(lessons)*100)+'%'
             }
        if course.link==None:
            inf.update({'link':'Ссылка на гугл диск пока не добавлена'})
        else:
            inf.update({'link': str(course.link)})
        for i in range(len(lessons)):
            inf.update({f'lesson_{i}': {'lesson_name':str(lessons[i].topic),
                                       'location':str(lessons[i].location),
                                       'date':str(lessons[i].date)
                                      }})
        return inf